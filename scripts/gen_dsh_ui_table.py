# -*- coding: utf-8 -*-
"""生成 dsh UI 布局意见填写表（Excel）"""
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

# ============ Sheet 1: UI 布局总览（合并单元格画三栏） ============
ws = wb.active
ws.title = "UI布局总览"

title_font = Font(bold=True, size=14, color="FFFFFF")
title_fill = PatternFill("solid", fgColor="1F4E78")
region_fill = {"sidebar": PatternFill("solid", fgColor="DDEBF7"),
               "center": PatternFill("solid", fgColor="E2EFDA"),
               "details": PatternFill("solid", fgColor="FCE4D6"),
               "overlay": PatternFill("solid", fgColor="FFF2CC")}
thin = Side(style="thin", color="999999")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
wrap = Alignment(wrap_text=True, vertical="top", horizontal="left")
center_a = Alignment(wrap_text=True, vertical="center", horizontal="center")

# 标题行
ws.merge_cells("A1:F1")
c = ws["A1"]; c.value = "DeepSeek Harness UI 布局总览（三栏结构）—— 在对应格子旁边填你的布局意见"
c.font = title_font; c.fill = title_fill; c.alignment = center_a
ws.row_dimensions[1].height = 28

# 列宽
for col, w in zip("ABCDEF", [18, 26, 30, 30, 26, 18]):
    ws.column_dimensions[col].width = w

def cell(addr, val, fill=None, font=None, align=wrap):
    c = ws[addr]; c.value = val
    c.alignment = align; c.border = border
    if fill: c.fill = fill
    if font: c.font = font
    return c

# 第一行：三栏标题
ws.merge_cells("A2:B2"); cell("A2", "① 侧栏 Sidebar\n（280px，可拖 264-420，<1024px 收成 56px 图标栏）", region_fill["sidebar"], Font(bold=True))
ws.merge_cells("C2:D2"); cell("C2", "② 中心区 Conversation\n（≥640px）", region_fill["center"], Font(bold=True))
ws.merge_cells("E2:F2"); cell("E2", "③ 详情面板 Details\n（360px，可拖 300-520）", region_fill["details"], Font(bold=True))
ws.row_dimensions[2].height = 40

# 行3-9: 侧栏 4 格 / 中心区 4 格 / 详情 1 格
sidebar_rows = [
    ("顶部", "品牌标识 Brand + 折叠开关", "sidebar（layout 拥有）"),
    ("次顶", "New Session 新建会话按钮", "sidebar（layout 拥有）"),
    ("中部", "工作区/会话浏览器：分区头+搜索+会话列表", "sidebar.workspaces（ui-workspace）"),
    ("底部", "设置入口 + 可选动作", "sidebar.settings / sidebar.footer.action"),
]
center_rows = [
    ("顶部条", "会话标题 + 视图标签页(chat/trajectory…) + 操作行", "conversation.session.header"),
    ("消息流", "消息节点流（assistant/command/tool/compaction/retry/error/turn-tail…）", "conversation.chat.node"),
    ("队列+输入", "队列坞 QueueDock + 输入栏 InputBar（附件/上下文计量/审批/待办）", "conversation（input/composer）"),
    ("空状态", "EmptyHero 引导页（无会话时）", "conversation.session"),
]
details_rows = [
    ("右侧面板", "详情视图（轨迹/工作流/目标/子代理/技能/交付物…随标签切换）", "details"),
]

r = 3
for i in range(4):
    label, content, slot = sidebar_rows[i]
    cl, cc, cs = f"A{r}", f"B{r}", f"C{r}"
    # 侧栏两列
    cell(f"A{r}", f"[侧栏·{label}] {content}", region_fill["sidebar"])
    cell(f"B{r}", f"插槽: {slot}", None, Font(italic=True, size=9))
    # 中心区两列
    if i < 4:
        label2, content2, slot2 = center_rows[i]
        cell(f"C{r}", f"[中心·{label2}] {content2}", region_fill["center"])
        cell(f"D{r}", f"插槽: {slot2}", None, Font(italic=True, size=9))
    # 详情两列（第一行占满，其余合并）
    if i == 0:
        cell(f"E{r}", f"[详情·{details_rows[0][0]}] {details_rows[0][1]}", region_fill["details"])
        cell(f"F{r}", f"插槽: {details_rows[0][2]}", None, Font(italic=True, size=9))
    else:
        ws.merge_cells(f"E{r}:F{r}")
        cell(f"E{r}", "（详情面板跨多行，内容随中心区标签切换）", region_fill["details"])
    ws.row_dimensions[r].height = 44
    r += 1

# 覆盖层行
ws.merge_cells("A10:F10")
cell("A10", "全局覆盖层 shell.overlay：对话框/模态/设置弹窗（模型选择、设置面板等）", region_fill["overlay"])
ws.row_dimensions[10].height = 28

# 图例行
ws.merge_cells("A11:F11")
cell("A11", "颜色说明：蓝=侧栏  绿=中心区  橙=详情面板  黄=全局覆盖层。填意见请到「布局意见表」Sheet，按区域编号填写。", None)
ws.row_dimensions[11].height = 24

# ============ Sheet 2: 布局意见表 ============
ws2 = wb.create_sheet("布局意见表")
headers = ["编号", "UI 区域", "插槽名", "现有布局说明", "你的布局意见（在此填写）", "改动类型", "备注"]
for i, h in enumerate(headers, 1):
    c = ws2.cell(row=1, column=i, value=h)
    c.font = Font(bold=True, color="FFFFFF")
    c.fill = title_fill
    c.alignment = center_a
    c.border = border
for col, w in zip("ABCDEFG", [8, 22, 30, 42, 46, 14, 30]):
    ws2.column_dimensions[col].width = w
ws2.row_dimensions[1].height = 24
ws2.freeze_panes = "A2"

rows = [
    ("S1", "侧栏·顶部品牌区", "sidebar", "品牌标识 + 折叠开关"),
    ("S2", "侧栏·新建会话", "sidebar", "New Session 按钮"),
    ("S3", "侧栏·会话浏览器", "sidebar.workspaces", "分区头 + 搜索 + 会话列表（分组/平铺）"),
    ("S4", "侧栏·底部设置区", "sidebar.settings / sidebar.footer.action", "设置入口 + 可选动作"),
    ("C1", "中心·顶部条", "conversation.session.header", "会话标题 + 视图标签页 + 操作行"),
    ("C2", "中心·消息流", "conversation.chat.node", "各类消息节点渲染（assistant/command/tool/compaction…）"),
    ("C3", "中心·输入区", "conversation.input", "输入栏 + 队列坞 + 附件 + 上下文计量 + 审批/待办面板"),
    ("C4", "中心·空状态", "conversation.session", "EmptyHero 引导页"),
    ("D1", "详情面板", "details", "详情视图（轨迹/工作流/目标/子代理/技能/交付物…）"),
    ("O1", "全局覆盖层", "shell.overlay", "模态框/设置弹窗/模型选择"),
    ("V1", "视图标签页", "conversation.view", "chat/trajectory 等标签（可新增标签）"),
    ("N1", "消息节点·思考过程", "conversation.chat.node（reasoning）", "ReasoningRow 思考过程展示"),
    ("N2", "消息节点·工具调用", "conversation.chat.node（tool）", "ToolCallBlock 工具调用块"),
    ("N3", "消息节点·统计", "conversation.chat.node（turn-metrics）", "StatsLine 统计行"),
]

row = 2
for rid, area, slot, desc in rows:
    ws2.cell(row=row, column=1, value=rid).border = border
    ws2.cell(row=row, column=2, value=area).border = border
    ws2.cell(row=row, column=3, value=slot).border = border
    ws2.cell(row=row, column=4, value=desc).border = border
    c5 = ws2.cell(row=row, column=5)
    c5.border = border; c5.fill = PatternFill("solid", fgColor="FFF9E6")
    ws2.cell(row=row, column=6, value="").border = border
    ws2.cell(row=row, column=7, value="").border = border
    for col in range(1, 8):
        ws2.cell(row=row, column=col).alignment = wrap
    ws2.row_dimensions[row].height = 40
    row += 1

# 底部说明
ws2.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
c = ws2.cell(row=row, column=1)
c.value = "用法：① 在「UI布局总览」看每个 UI 位置 → ② 在「布局意见表」对应编号行的黄色格写你的意见 → ③ 改动类型选：改样式/改结构/加内容/删除/新增区域 → ④ 保存后发我，我按你的意见评估每个插槽怎么改。"
c.alignment = wrap
ws2.row_dimensions[row].height = 42

out = r"C:\Users\23001\Desktop\dsh_UI布局意见表.xlsx"
try:
    wb.save(out)
    print("已保存:", out)
except Exception as e:
    out2 = r"C:\Users\23001\AppData\Roaming\reasonix\global-workspace\dsh_UI布局意见表.xlsx"
    wb.save(out2)
    print("桌面不可写，已保存到:", out2)
